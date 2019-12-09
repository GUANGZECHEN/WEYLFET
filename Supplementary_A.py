# -*- coding: utf-8 -*-
"""
Created on Tue Aug  6 22:45:00 2019

@author: Guangze Chen
"""
# This script is to calculate bandstructure of the rotated Weyl semimetal nanowire and the conductance along the nanowire for dispersive Fermi arcs in presence of disorder
# Figs.3(a) and (b) and Fig.4 are based on results of this script
# To run this code, type "python sigma_z_disorder.py i j" where i and j corresponds to the tilting angle theta and the gate voltage strength/disorder strength, see below 

from math import sin, cos, sqrt, pi
import kwant
from matplotlib import pyplot
import sys
from openpyxl import Workbook
from kwant.digest import gauss

# In the following codes, we have 10^(-10)m=1 and 10^(-20)J=1

# The following function is to construct the system using the lattice model we obtained for the Weyl semimetal
def make_system(phi,V,V_disorder): 
    # phi: the angle we rotate the Weyl points, phi=pi/2 corresponds to the case of theta=0 in the article; 
    # V: gate voltage added on surface to introduce dispersive Fermi arcs; 
    # V_disorder: strength of disorder
    syst = kwant.Builder()
    lat_u = kwant.lattice.cubic(a=100,name='u')
    lat_d = kwant.lattice.cubic(a=100,name='d')
  
    a=100 # lattice constant, a=10nm
    W=40  # width of system in units of lattice constant
    L=40  # length of scattering region, where disorder can be added
    x,y,z=W,W,L
    M=0.7 # parameter M=7*10^(-21)J, which means the M_minimal in the minimal model is M_minimal=Ma^2=7*10^(-19)J*nm^2=4.375eV*nm^2
    k_0=1 # k_0=0.1*nm^(-1)
    U=V   # gate voltage added on surface
    U2=-U # This establish the system with U_I=-U_II, which corresponds to figs.4(a) and (c). To get results for U_I=U_II in figs.4(b) and (d), set U2=U 
    alpha=V_disorder

    def disorder(site,salt): # generate a Gaussian type disorder
        return alpha*gauss(repr(site),salt)

    for i in range(x):
        for j in range(y):
            for k in range(z):
                syst[lat_u(i,j,k)] = M*(k_0**2-6)
                syst[lat_d(i,j,k)] = -M*(k_0**2-6)
                
                if i>0:
                    syst[lat_u(i,j,k),lat_u(i-1,j,k)] = M
                    syst[lat_d(i,j,k),lat_d(i-1,j,k)] = -M
                    syst[lat_u(i,j,k),lat_d(i-1,j,k)] = 1j/2*(cos(phi/2)**2)+1/2*(sin(phi/2)**2)
                    syst[lat_d(i,j,k),lat_u(i-1,j,k)] = 1j/2*(cos(phi/2)**2)-1/2*(sin(phi/2)**2)
                        
                if j>0:
                    syst[lat_u(i,j,k),lat_u(i,j-1,k)] = M
                    syst[lat_d(i,j,k),lat_d(i,j-1,k)] = -M
                    syst[lat_u(i,j,k),lat_d(i,j-1,k)] = 1/2*(cos(phi/2)**2)+1j/2*(sin(phi/2)**2)
                    syst[lat_d(i,j,k),lat_u(i,j-1,k)] = -1/2*(cos(phi/2)**2)+1j/2*(sin(phi/2)**2)
                        
                if k>0:
                    syst[lat_u(i,j,k),lat_u(i,j,k-1)] = M
                    syst[lat_d(i,j,k),lat_d(i,j,k-1)] = -M
                    syst[lat_u(i,j,k),lat_d(i,j,k-1)] = 1/(sqrt(8))*(1-1j)*sin(phi)
                    syst[lat_d(i,j,k),lat_u(i,j,k-1)] = 1/(sqrt(8))*(-1-1j)*sin(phi)
    
    for i in range(x):
        for k in range(z):
            V1=disorder((i,0,k),"=.=")              # "=.=" represent salt of the disorder, and is arbitrary, e.g. "", "+", " "
            syst[lat_u(i,0,k)] = M*(k_0**2-6)+U+V1
            syst[lat_d(i,0,k)] = -M*(k_0**2-6)+U+V1
            
            V2=disorder((i,y-1,k),"=.=")
            syst[lat_u(i,y-1,k)] = M*(k_0**2-6)+U+V2
            syst[lat_d(i,y-1,k)] = -M*(k_0**2-6)+U+V2
    
    for j in range(y):
        for k in range(z):
            V3=disorder((0,j,k),"=.=")
            syst[lat_u(0,j,k)] = M*(k_0**2-6)+U2+V3
            syst[lat_d(0,j,k)] = -M*(k_0**2-6)+U2+V3
            
            V4=disorder((x-1,j,k),"=.=")
            syst[lat_u(x-1,j,k)] = M*(k_0**2-6)+U2+V4
            syst[lat_d(x-1,j,k)] = -M*(k_0**2-6)+U2+V4
    
    ### Define and attach the leads. ####
    lead = kwant.Builder(kwant.TranslationalSymmetry((0,0,-a)))
  
    for i in range(x):
        for j in range(y):
            lead[lat_u(i,j,0)] = M*(k_0**2-6)
            lead[lat_d(i,j,0)] = -M*(k_0**2-6)
            if i>0:
                lead[lat_u(i,j,0),lat_u(i-1,j,0)] = M
                lead[lat_d(i,j,0),lat_d(i-1,j,0)] = -M
                lead[lat_u(i,j,0),lat_d(i-1,j,0)] = 1j/2*(cos(phi/2)**2)+1/2*(sin(phi/2)**2)
                lead[lat_d(i,j,0),lat_u(i-1,j,0)] = 1j/2*(cos(phi/2)**2)-1/2*(sin(phi/2)**2)
                        
            if j>0:
                lead[lat_u(i,j,0),lat_u(i,j-1,0)] = M
                lead[lat_d(i,j,0),lat_d(i,j-1,0)] = -M
                lead[lat_u(i,j,0),lat_d(i,j-1,0)] = 1/2*(cos(phi/2)**2)+1j/2*(sin(phi/2)**2)
                lead[lat_d(i,j,0),lat_u(i,j-1,0)] = -1/2*(cos(phi/2)**2)+1j/2*(sin(phi/2)**2)
            
            lead[lat_u(i,j,0),lat_u(i,j,-1)] = M
            lead[lat_d(i,j,0),lat_d(i,j,-1)] = -M
            lead[lat_u(i,j,0),lat_d(i,j,-1)] = 1/(sqrt(8))*(1-1j)*sin(phi)
            lead[lat_d(i,j,0),lat_u(i,j,-1)] = -1/(sqrt(8))*(1+1j)*sin(phi)
            
    for i in range(x):
        lead[lat_u(i,0,0)] = M*(k_0**2-6)+U
        lead[lat_d(i,0,0)] = -M*(k_0**2-6)+U
        lead[lat_u(i,y-1,0)] = M*(k_0**2-6)+U
        lead[lat_d(i,y-1,0)] = -M*(k_0**2-6)+U
        
    for j in range(y):
        lead[lat_u(0,j,0)] = M*(k_0**2-6)+U2
        lead[lat_d(0,j,0)] = -M*(k_0**2-6)+U2
        lead[lat_u(x-1,j,0)] = M*(k_0**2-6)+U2
        lead[lat_d(x-1,j,0)] = -M*(k_0**2-6)+U2
            
    lead2 = kwant.Builder(kwant.TranslationalSymmetry((0,0,a)))
  
    for i in range(x):
        for j in range(y):
            lead2[lat_u(i,j,L-1)] = M*(k_0**2-6)
            lead2[lat_d(i,j,L-1)] = -M*(k_0**2-6)
            if i>0:
                lead2[lat_u(i,j,L-1),lat_u(i-1,j,L-1)] = M
                lead2[lat_d(i,j,L-1),lat_d(i-1,j,L-1)] = -M
                lead2[lat_u(i,j,L-1),lat_d(i-1,j,L-1)] = 1j/2*(cos(phi/2)**2)+1/2*(sin(phi/2)**2)
                lead2[lat_d(i,j,L-1),lat_u(i-1,j,L-1)] = 1j/2*(cos(phi/2)**2)-1/2*(sin(phi/2)**2)
                        
            if j>0:
                lead2[lat_u(i,j,L-1),lat_u(i,j-1,L-1)] = M
                lead2[lat_d(i,j,L-1),lat_d(i,j-1,L-1)] = -M
                lead2[lat_u(i,j,L-1),lat_d(i,j-1,L-1)] = 1/2*(cos(phi/2)**2)+1j/2*(sin(phi/2)**2)
                lead2[lat_d(i,j,L-1),lat_u(i,j-1,L-1)] = -1/2*(cos(phi/2)**2)+1j/2*(sin(phi/2)**2)
            
            lead2[lat_u(i,j,L),lat_u(i,j,L-1)] = M
            lead2[lat_d(i,j,L),lat_d(i,j,L-1)] = -M
            lead2[lat_u(i,j,L),lat_d(i,j,L-1)] = 1/(sqrt(8))*(1-1j)*sin(phi)
            lead2[lat_d(i,j,L),lat_u(i,j,L-1)] = -1/(sqrt(8))*(1+1j)*sin(phi)
   
    for i in range(x):
        lead2[lat_u(i,0,z-1)] = M*(k_0**2-6)+U
        lead2[lat_d(i,0,z-1)] = -M*(k_0**2-6)+U
        lead2[lat_u(i,y-1,z-1)] = M*(k_0**2-6)+U
        lead2[lat_d(i,y-1,z-1)] = -M*(k_0**2-6)+U
        
    for j in range(y):
        lead2[lat_u(0,j,z-1)] = M*(k_0**2-6)+U2
        lead2[lat_d(0,j,z-1)] = -M*(k_0**2-6)+U2
        lead2[lat_u(x-1,j,z-1)] = M*(k_0**2-6)+U2
        lead2[lat_d(x-1,j,z-1)] = -M*(k_0**2-6)+U2
        
    syst.attach_lead(lead)
    syst.attach_lead(lead2)
    
    syst1=syst.finalized()
    
    return syst1

def bandstructure(syst):
    flead=syst.leads[0]
    bands_0=kwant.physics.Bands(flead)
    momenta = [-0.2 + 0.004 * i for i in range(101)]
    energies = [bands_0(k) for k in momenta]
    
    pyplot.figure()
    pyplot.plot(momenta, energies,color='0',linewidth=0.4)
    pyplot.axis([-0.2,0.2,-1,1])
    pyplot.xlabel("kz")
    pyplot.ylabel("E")
    pyplot.show()
    
def compute_conductance(i,j):
    # Compute conductance
    wb=Workbook()
    sh=wb.active
    sh.title="ads"    
    
    # Here we calculate the conductance in presence of disorder for fixed surface gate voltage V=0.2=12.5meV, which corresponds to figs.4(c) and (d)
    # With a small change we can calculate the conductance with different gate voltage in absence of disorder
    V=0.2
    phi=pi/2-pi/300*j
    V_disorder=0.2*i
    syst = make_system(phi,V,V_disorder)
    smatrix = kwant.smatrix(syst, 0.0001) # To avoid numerical singularities, the bias is set to 0.0001=6.25*10^(-3)meV, this bias is much smaller than the bulk gap for this system (even at theta=0 the bulk is gapped since this is a finite system) and does not introduce bulk transport
    sh.cell(row=j+1,column=1).value=pi/2-phi
    sh.cell(row=j+1,column=i+2).value=smatrix.transmission(1,0)
    wb.save("2_node_Weyl"+str(i)+"_"+str(j)+".xlsx") # store the data for further use

def main():
    i=sys.argv[1]
    i=int(i)
    j=sys.argv[2]
    j=int(j)
    compute_conductance(i,j)
   
# run this function to get bandstructure for the system
def main2():
    syst=make_system(pi/2,0,0) # with phi=pi/2 we get fig3(a) and phi=pi/3 we get fig3(b)
    kwant.plot(syst)
    bandstructure(syst)

# Call the main function if the script gets executed (as opposed to imported).
# See <http://docs.python.org/library/__main__.html>.
if __name__ == '__main__':
    main()
