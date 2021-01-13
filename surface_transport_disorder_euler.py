# -*- coding: utf-8 -*-
"""
Created on Fri Oct 19 18:06:59 2018

@author: cgz
"""

# Tutorial 2.2.3. Building the same system with less code
# =======================================================
#
# Physics background
# ------------------
#  Conductance of a quantum wire; subbands
#
# Kwant features highlighted
# --------------------------
#  - Using iterables and builder.HoppingKind for making systems
#  - introducing `reversed()` for the leads
#
# Note: Does the same as tutorial1a.py, but using other features of Kwant.
from __future__ import division
from math import sin, cos, sqrt, pi, acos
import kwant
#import tinyarray
#import scipy.sparse.linalg as sla
# For plotting
from matplotlib import pyplot
import sys
#import xlwt
from openpyxl import Workbook
from kwant.digest import gauss
#from openpyxl import load_workbook

#length scale: A=10^-10m

def make_system(pos,disorder,theta,W,L):
    syst = kwant.Builder()
    lat_u = kwant.lattice.cubic(a=50,name='u',norbs=1)
    lat_d = kwant.lattice.cubic(a=50,name='d',norbs=1)
    lat_m = kwant.lattice.cubic(a=50,name='m',norbs=1)
  
    a=50
    x,y,z=W,W,L
  #  z1,z2=0,1
   # z3,z4=L-1,L
    x1,x2=19,20
    z5,z6=19,20
    M=-20
    k_0=0.5
    D=10
    t=0.01
    t2=1
    v=2
   # E_0=0
    alpha=0.05*disorder-0.2
    
    def disorder1(site,salt):
        return alpha*gauss(repr(site),salt) + (M*(k_0**2-6)+D)*v*0.02
    
    def disorder2(site,salt):
        return alpha*gauss(repr(site),salt) + (-M*(k_0**2-6)+D)*v*0.02
    
    def disorder3(site,salt):
        return alpha*gauss(repr(site),salt) + (M*(k_0**2-6)-D)*v*0.02
    
    def disorder4(site,salt):
        return alpha*gauss(repr(site),salt) + (-M*(k_0**2-6)-D)*v*0.02
    
    for i in range(x):
        for j in range(y):
            for k in range(z):
                syst[lat_u(i,j,k)] = M*(k_0**2-6)*v*0.02
                syst[lat_d(i,j,k)] = -M*(k_0**2-6)*v*0.02
                
                if i>0:
                    syst[lat_u(i,j,k),lat_u(i-1,j,k)] = M*v*0.02
                    syst[lat_d(i,j,k),lat_d(i-1,j,k)] = -M*v*0.02
                    syst[lat_u(i,j,k),lat_d(i-1,j,k)] = (1j/2*(cos(theta/2)**2)+1/2*(sin(theta/2)**2))*v
                    syst[lat_d(i,j,k),lat_u(i-1,j,k)] = (1j/2*(cos(theta/2)**2)-1/2*(sin(theta/2)**2))*v
                        
                if j>0:
                    syst[lat_u(i,j,k),lat_u(i,j-1,k)] = M*v*0.02
                    syst[lat_d(i,j,k),lat_d(i,j-1,k)] = -M*v*0.02
                    syst[lat_u(i,j,k),lat_d(i,j-1,k)] = (1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
                    syst[lat_d(i,j,k),lat_u(i,j-1,k)] = (-1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v

                if k>0:
                    syst[lat_u(i,j,k),lat_u(i,j,k-1)] = M*v*0.02
                    syst[lat_d(i,j,k),lat_d(i,j,k-1)] = -M*v*0.02
                    syst[lat_u(i,j,k),lat_d(i,j,k-1)] = 1/(sqrt(8))*(1-1j)*sin(theta)*v
                    syst[lat_d(i,j,k),lat_u(i,j,k-1)] = -1/(sqrt(8))*(1+1j)*sin(theta)*v
                    
    #syst[(lat_u(i,y-1,k) for i in range(x) for k in range(z))] = disorder1
    #syst[(lat_d(i,y-1,k) for i in range(x) for k in range(z))] = disorder2
    #syst[(lat_u(i,0,k) for i in range(x) for k in range(z))] = disorder1
    #syst[(lat_d(i,0,k) for i in range(x) for k in range(z))] = disorder2
    #syst[(lat_u(x-1,j,k) for j in range(y) for k in range(z))] = disorder3
    #syst[(lat_d(x-1,j,k) for j in range(y) for k in range(z))] = disorder4
    #syst[(lat_u(0,j,k) for j in range(y) for k in range(z))] = disorder3
    #syst[(lat_d(0,j,k) for j in range(y) for k in range(z))] = disorder4
    
    for i in range(x):
        for k in range(z):
            U=disorder1((i,0,k),"=.=") - (M*(k_0**2-6)+D)*v*0.02
            syst[lat_u(i,0,k)]=U+(M*(k_0**2-6)+D)*v*0.02
            syst[lat_d(i,0,k)]=U+(-M*(k_0**2-6)+D)*v*0.02
    
    for j in range(y):
        for k in range(z):
            U2=disorder3((x-1,j,k),"=.=") - (M*(k_0**2-6)-D)*v*0.02
            syst[lat_u(x-1,j,k)]=U2+(M*(k_0**2-6)-D)*v*0.02
            syst[lat_d(x-1,j,k)]=U2+(-M*(k_0**2-6)-D)*v*0.02
            
#final lead at x+            
    for j in range(pos,pos+1):
        for k in range(z5,z6):
            syst[lat_m(x,j,k)]=0
            syst[lat_m(x,j,k),lat_u(x-1,j,k)]=t
            syst[lat_m(x,j,k),lat_d(x-1,j,k)]=t
            if j>pos:
                syst[lat_m(x,j,k),lat_m(x,j-1,k)]=t2
            if k>z5:
                syst[lat_m(x,j,k),lat_m(x,j,k-1)]=t2

    lead = kwant.Builder(kwant.TranslationalSymmetry((a,0,0)))
  
    for j in range(pos,pos+1):
        for k in range(z5,z6):
            lead[lat_m(x,j,k)]=0
            lead[lat_m(x,j,k),lat_m(x+1,j,k)]=t2
            if j>pos:
                lead[lat_m(x,j,k),lat_m(x,j,k)]=t2
            if k>z5:
                lead[lat_m(x,j,k),lat_m(x,j,k-1)]=t2        
   
#making z direction infinite    
    lead3 = kwant.Builder(kwant.TranslationalSymmetry((0,0,-a)))
  
    for i in range(x):
        for j in range(y):
            lead3[lat_u(i,j,0)] = M*(k_0**2-6)*v*0.02
            lead3[lat_d(i,j,0)] = -M*(k_0**2-6)*v*0.02
            if i>0:
                lead3[lat_u(i,j,0),lat_u(i-1,j,0)] = M*v*0.02
                lead3[lat_d(i,j,0),lat_d(i-1,j,0)] = -M*v*0.02
                lead3[lat_u(i,j,0),lat_d(i-1,j,0)] = (1j/2*(cos(theta/2)**2)+1/2*(sin(theta/2)**2))*v
                lead3[lat_d(i,j,0),lat_u(i-1,j,0)] = (1j/2*(cos(theta/2)**2)-1/2*(sin(theta/2)**2))*v
                        
            if j>0:
                lead3[lat_u(i,j,0),lat_u(i,j-1,0)] = M*v*0.02
                lead3[lat_d(i,j,0),lat_d(i,j-1,0)] = -M*v*0.02
                lead3[lat_u(i,j,0),lat_d(i,j-1,0)] = (1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
                lead3[lat_d(i,j,0),lat_u(i,j-1,0)] = (-1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
            
            lead3[lat_u(i,j,0),lat_u(i,j,-1)] = M*v*0.02
            lead3[lat_d(i,j,0),lat_d(i,j,-1)] = -M*v*0.02
            lead3[lat_u(i,j,0),lat_d(i,j,-1)] = 1/(sqrt(8))*(1-1j)*sin(theta)*v
            lead3[lat_d(i,j,0),lat_u(i,j,-1)] = -1/(sqrt(8))*(1+1j)*sin(theta)*v
            
    for i in range(x):
        lead3[lat_u(i,0,0)]=(M*(k_0**2-6)+D)*v*0.02
        lead3[lat_d(i,0,0)]=(-M*(k_0**2-6)+D)*v*0.02
    
    for j in range(y):
        lead3[lat_u(x-1,j,0)]=(M*(k_0**2-6)-D)*v*0.02
        lead3[lat_d(x-1,j,0)]=(-M*(k_0**2-6)-D)*v*0.02
            
    lead4 = kwant.Builder(kwant.TranslationalSymmetry((0,0,a)))
  
    for i in range(x):
        for j in range(y):
            lead4[lat_u(i,j,L-1)] = (M*(k_0**2-6))*v*0.02
            lead4[lat_d(i,j,L-1)] = (-M*(k_0**2-6))*v*0.02
            if i>0:
                lead4[lat_u(i,j,L-1),lat_u(i-1,j,L-1)] = M*v*0.02
                lead4[lat_d(i,j,L-1),lat_d(i-1,j,L-1)] = -M*v*0.02
                lead4[lat_u(i,j,L-1),lat_d(i-1,j,L-1)] = (1j/2*(cos(theta/2)**2)+1/2*(sin(theta/2)**2))*v
                lead4[lat_d(i,j,L-1),lat_u(i-1,j,L-1)] = (1j/2*(cos(theta/2)**2)-1/2*(sin(theta/2)**2))*v
                        
            if j>0:
                lead4[lat_u(i,j,L),lat_u(i,j-1,L)] = M*v*0.02
                lead4[lat_d(i,j,L),lat_d(i,j-1,L)] = -M*v*0.02
                lead4[lat_u(i,j,L),lat_d(i,j-1,L)] = (1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
                lead4[lat_d(i,j,L),lat_u(i,j-1,L)] = (-1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v

            lead4[lat_u(i,j,L),lat_u(i,j,L-1)] = M*v*0.02
            lead4[lat_d(i,j,L),lat_d(i,j,L-1)] = -M*v*0.02
            lead4[lat_u(i,j,L),lat_d(i,j,L-1)] = 1/(sqrt(8))*(1-1j)*sin(theta)*v
            lead4[lat_d(i,j,L),lat_u(i,j,L-1)] = -1/(sqrt(8))*(1+1j)*sin(theta)*v
            
    for i in range(x):
        lead4[lat_u(i,0,z-1)]=(M*(k_0**2-6)+D)*v*0.02
        lead4[lat_d(i,0,z-1)]=(-M*(k_0**2-6)+D)*v*0.02
    
    for j in range(y):
        lead4[lat_u(x-1,j,z-1)]=(M*(k_0**2-6)-D)*v*0.02
        lead4[lat_d(x-1,j,z-1)]=(-M*(k_0**2-6)-D)*v*0.02
        
    for i in range(x1,x2):
        for k in range(z5,z6):
            syst[lat_m(i,-1,k)]=0
            syst[lat_m(i,-1,k),lat_u(i,0,k)]=t
            syst[lat_m(i,-1,k),lat_d(i,0,k)]=t
            if i>x1:
                syst[lat_m(i,-1,k),lat_m(i-1,-1,k)]=t2
            if k>z5:
                syst[lat_m(i,-1,k),lat_m(i,-1,k-1)]=t2
                
    lead5 = kwant.Builder(kwant.TranslationalSymmetry((0,-a,0)))
  
    for i in range(x1,x2):
        for k in range(z5,z6):
            lead5[lat_m(i,-1,k)]=0
            lead5[lat_m(i,-1,k),lat_m(i,-2,k)]=t2
            if i>x1:
                lead5[lat_m(i,-1,k),lat_m(i-1,-1,k)]=t2
            if k>z5:
                lead5[lat_m(i,-1,k),lat_m(i,-1,k-1)]=t2
    
    lead6 = kwant.Builder(kwant.TranslationalSymmetry((0,a,0)))
    
    for i in range(x):
        for k in range(z):
            lead6[lat_u(i,y,k)] = (M*(k_0**2-6))*v*0.02
            lead6[lat_d(i,y,k)] = -(M*(k_0**2-6))*v*0.02
            lead6[lat_u(i,y-1,k)] = (M*(k_0**2-6))*v*0.02
            lead6[lat_d(i,y-1,k)] = -(M*(k_0**2-6))*v*0.02
     
    for i in range(x):
        for k in range(z):   
            if i>0:
                lead6[lat_u(i,y,k),lat_u(i-1,y,k)] = M*v*0.02
                lead6[lat_d(i,y,k),lat_d(i-1,y,k)] = -M*v*0.02
                lead6[lat_u(i,y,k),lat_d(i-1,y,k)] = (1j/2*(cos(theta/2)**2)+1/2*(sin(theta/2)**2))*v
                lead6[lat_d(i,y,k),lat_u(i-1,y,k)] = (1j/2*(cos(theta/2)**2)-1/2*(sin(theta/2)**2))*v
    
            lead6[lat_u(i,y,k),lat_u(i,y-1,k)] = M*v*0.02
            lead6[lat_d(i,y,k),lat_d(i,y-1,k)] = -M*v*0.02
            lead6[lat_u(i,y,k),lat_d(i,y-1,k)] = (1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
            lead6[lat_d(i,y,k),lat_u(i,y-1,k)] = (-1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
            
            if k>0:                
                lead6[lat_u(i,y,k),lat_u(i,y,k-1)] = M*v*0.02
                lead6[lat_d(i,y,k),lat_d(i,y,k-1)] = -M*v*0.02
                lead6[lat_u(i,y,k),lat_d(i,y,k-1)] = 1/(sqrt(8))*(1-1j)*sin(theta)*v
                lead6[lat_d(i,y,k),lat_u(i,y,k-1)] = -1/(sqrt(8))*(1+1j)*sin(theta)*v
                
    for k in range(z):
        lead6[lat_u(x-1,y-1,k)]=(M*(k_0**2-6)-D)*v*0.02
        lead6[lat_d(x-1,y-1,k)]=(-M*(k_0**2-6)-D)*v*0.02

                
    lead7 = kwant.Builder(kwant.TranslationalSymmetry((-a,0,0)))
    
    for j in range(y):
        for k in range(z):
            lead7[lat_u(0,j,k)] = (M*(k_0**2-6))*v*0.02
            lead7[lat_d(0,j,k)] = -(M*(k_0**2-6))*v*0.02          
            lead7[lat_u(-1,j,k)] = (M*(k_0**2-6))*v*0.02
            lead7[lat_d(-1,j,k)] = -(M*(k_0**2-6))*v*0.02
  
            
    for j in range(y):
        for k in range(z):
            lead7[lat_u(0,j,k),lat_u(-1,j,k)] = M*v*0.02
            lead7[lat_d(0,j,k),lat_d(-1,j,k)] = -M*v*0.02
            lead7[lat_u(0,j,k),lat_d(-1,j,k)] = (1j/2*(cos(theta/2)**2)+1/2*(sin(theta/2)**2))*v
            lead7[lat_d(0,j,k),lat_u(-1,j,k)] = (1j/2*(cos(theta/2)**2)-1/2*(sin(theta/2)**2))*v
            
            if j>0:                
                lead7[lat_u(0,j,k),lat_u(0,j-1,k)] = M*v*0.02
                lead7[lat_d(0,j,k),lat_d(0,j-1,k)] = -M*v*0.02
                lead7[lat_u(0,j,k),lat_d(0,j-1,k)] = (1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
                lead7[lat_d(0,j,k),lat_u(0,j-1,k)] = (-1/2*(cos(theta/2)**2)+1j/2*(sin(theta/2)**2))*v
                
            if k>0:                
                lead7[lat_u(0,j,k),lat_u(0,j,k-1)] = M*v*0.02
                lead7[lat_d(0,j,k),lat_d(0,j,k-1)] = -M*v*0.02
                lead7[lat_u(0,j,k),lat_d(0,j,k-1)] = 1/(sqrt(8))*(1-1j)*sin(theta)*v
                lead7[lat_d(0,j,k),lat_u(0,j,k-1)] = -1/(sqrt(8))*(1+1j)*sin(theta)*v
                
    for k in range(z):
        lead7[lat_u(0,0,k)]=(M*(k_0**2-6)+D)*v*0.02
        lead7[lat_d(0,0,k)]=(-M*(k_0**2-6)+D)*v*0.02
                    
    # Attach the left lead and its reversed copy.
    syst.attach_lead(lead)
    syst.attach_lead(lead3)
    syst.attach_lead(lead4)
    syst.attach_lead(lead5)
    syst.attach_lead(lead6)
    syst.attach_lead(lead7)
    
    syst1=syst.finalized()
    
    return syst1

def plot_conductance(i,j):
    # Compute conductance
    wb=Workbook()
    sh=wb.active
    sh.title="ads"    

    syst = make_system(i,j,acos(1/sqrt(3)),40,40)
    smatrix = kwant.smatrix(syst, 0, args=[""])
    sh.cell(row=i+1,column=1).value=i
    sh.cell(row=i+1,column=2).value=smatrix.transmission(3,0)
    sh.cell(row=i+1,column=3).value=smatrix.transmission(0,3)
    wb.save("2_node_Weyl"+str(i)+"_"+str(j)+".xlsx")

def bands(syst):
    #for i in range(1):
    flead=syst.leads[5]
    bands_0=kwant.physics.Bands(flead)
    momenta = [-3*pi/4 + 0.04 * 3*pi/4 * i for i in range(51)]
    energies = [bands_0(k) for k in momenta]
    
    pyplot.figure()
    pyplot.plot(momenta, energies)
    pyplot.axis([-3*pi/4,3*pi/4,-1,1])
    pyplot.xlabel("k_x")
    pyplot.ylabel("E") 
    pyplot.show()
    
def main():

  #  i=sys.argv[1]
   # j=sys.argv[2]
   # i=int(i)
   # j=int(j)
   # plot_conductance(i,j)
    syst=make_system(10,0,acos(1/sqrt(3)),30,30)
   # smatrix = kwant.smatrix(syst, 0, args=[""])
   # x=smatrix.transmission(3,0)
   # print(x)
   # syst2=make_system2(30)
   # kwant.plot(syst2)
   # wf = kwant.wave_function(syst, energy=0,args=[""])
   # psi=wf(0)[0]
   # def surface1(site_in,site_out):
   #     return site_in.pos[0] == 0 and site_out.pos[0] == 0
   # J_1 = kwant.operator.Current(syst,where=surface1)
   # def surface2(site_in,site_out):
   #     return site_in.pos[1] == 29 and site_out.pos[1] == 29
   # J_2 = kwant.operator.Current(syst,where=surface2)
   # def surface3(site_in,site_out):
   #     return site_in.pos[0] == 29 and site_out.pos[0] == 29
   # J_3 = kwant.operator.Current(syst,where=surface3)
   # def surface4(site_in,site_out):
   #     return site_in.pos[1] == 0 and site_out.pos[1] == 0
   # J_4 = kwant.operator.Current(syst,where=surface4)
   # current1 = J_1(psi)
   # current2 = J_2(psi)
   # current3 = J_3(psi)
   # current4 = J_4(psi)
    #print(density)
   # kwant.plotter.current(syst2, current1, colorbar=True)
   # kwant.plotter.current(syst2, current2, colorbar=True)
   # kwant.plotter.current(syst2, current3, colorbar=True)
   # kwant.plotter.current(syst2, current4, colorbar=True)
    kwant.plot(syst)
   # bands(syst)
   # plot_conductance_on_my_own()
# Call the main function if the script gets executed (as opposed to imported).
# See <http://docs.python.org/library/__main__.html>.
if __name__ == '__main__':
    main()
